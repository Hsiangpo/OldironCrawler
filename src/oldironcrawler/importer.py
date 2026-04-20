from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from urllib.parse import urlparse

from openpyxl import load_workbook


_SUPPORTED_SUFFIXES = {".txt", ".csv", ".xlsx"}
_HEADER_CANDIDATES = ("website", "url", "domain", "homepage")
_POSITIVE_HEADER_HINTS = {
    "website": 40,
    "company website": 60,
    "official website": 60,
    "company url": 55,
    "homepage": 45,
    "home page": 45,
    "domain": 35,
    "company domain": 40,
    "web": 10,
    "url": 8,
    "site": 6,
}
_NEGATIVE_HEADER_HINTS = {
    "linkedin": -120,
    "facebook": -120,
    "instagram": -120,
    "twitter": -120,
    "x.com": -120,
    "youtube": -120,
    "tiktok": -120,
    "social": -100,
    "email": -140,
    "e-mail": -140,
    "mail": -110,
    "phone": -90,
    "mobile": -90,
    "address": -90,
    "note": -80,
    "remark": -80,
    "comment": -80,
}
_SOCIAL_HOST_HINTS = (
    "linkedin.com",
    "facebook.com",
    "instagram.com",
    "twitter.com",
    "x.com",
    "youtube.com",
    "tiktok.com",
    "wechat.com",
    "weibo.com",
)
_WEBSITE_COLUMN_BONUS = {"high": 90.0, "medium": 60.0, "low": 30.0}


WebsiteColumnPicker = Callable[..., dict[str, object]]


@dataclass
class ImportedWebsite:
    input_index: int
    raw_website: str
    website: str
    dedupe_key: str


@dataclass
class WebsiteColumnSummary:
    index: int
    header: str
    sample_values: list[str]
    non_empty_count: int
    website_count: int
    homepage_count: int
    social_count: int
    email_count: int
    note_count: int
    local_score: float

    def to_llm_payload(self) -> dict[str, object]:
        return {
            "index": self.index,
            "header": self.header,
            "samples": self.sample_values,
            "non_empty_count": self.non_empty_count,
            "website_count": self.website_count,
            "homepage_count": self.homepage_count,
            "social_count": self.social_count,
            "email_count": self.email_count,
            "note_count": self.note_count,
            "local_score": round(self.local_score, 2),
        }


@dataclass
class WebsiteColumnSelection:
    column_index: int
    header: str
    confidence: str
    reason: str
    skip_header: bool


def list_input_files(websites_dir: Path) -> list[Path]:
    if not websites_dir.exists():
        return []
    items = [path for path in websites_dir.iterdir() if path.is_file() and path.suffix.lower() in _SUPPORTED_SUFFIXES]
    return sorted(items, key=lambda item: item.name.lower())


def choose_input_file(websites_dir: Path) -> Path:
    files = list_input_files(websites_dir)
    if not files:
        raise RuntimeError(f"没有找到输入文件，请把 txt/csv/xlsx 放到 {websites_dir}")
    name_map = {path.name.lower(): path for path in files}
    print("可用输入文件：")
    for index, path in enumerate(files, start=1):
        print(f"  {index}. {path.name}")
    while True:
        raw = input("请输入文件序号: ").strip()
        matched = name_map.get(raw.lower())
        if matched is not None:
            return matched
        try:
            choice = int(raw)
        except ValueError:
            print("输入无效，请重新输入序号或文件名。")
            continue
        if 1 <= choice <= len(files):
            return files[choice - 1]
        print("序号超出范围，请重新输入。")


def compute_file_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def compute_rows_fingerprint(rows: list[ImportedWebsite]) -> str:
    payload = [
        {
            "input_index": row.input_index,
            "website": row.website,
            "dedupe_key": row.dedupe_key,
        }
        for row in rows
    ]
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")).hexdigest()


def load_websites(
    path: Path,
    *,
    website_column_picker: WebsiteColumnPicker | None = None,
) -> list[ImportedWebsite]:
    suffix = path.suffix.lower()
    if suffix == ".txt":
        rows = _load_from_txt(path)
    elif suffix == ".csv":
        rows = _load_from_csv(path, website_column_picker=website_column_picker)
    elif suffix == ".xlsx":
        rows = _load_from_xlsx(path, website_column_picker=website_column_picker)
    else:
        raise RuntimeError(f"不支持的文件类型: {path.suffix}")
    return _dedupe_websites(rows)


def _load_from_txt(path: Path) -> list[str]:
    rows: list[str] = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        text = str(line or "").strip()
        if text:
            rows.append(text)
    return rows


def _load_from_csv(
    path: Path,
    *,
    website_column_picker: WebsiteColumnPicker | None = None,
) -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    return _load_from_matrix(rows, source_name=path.name, website_column_picker=website_column_picker)


def _load_from_xlsx(
    path: Path,
    *,
    website_column_picker: WebsiteColumnPicker | None = None,
) -> list[str]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        rows: list[str] = []
        for sheet in workbook.worksheets:
            matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
            if not matrix:
                continue
            rows.extend(
                _load_from_matrix(
                    matrix,
                    source_name=f"{path.name} | {sheet.title}",
                    website_column_picker=website_column_picker,
                )
            )
    finally:
        workbook.close()
    return rows


def _load_from_matrix(
    rows: list[list[object]],
    *,
    source_name: str = "",
    website_column_picker: WebsiteColumnPicker | None = None,
) -> list[str]:
    if not rows:
        return []
    selection = _pick_website_column(rows, source_name=source_name, website_column_picker=website_column_picker)
    if selection is not None:
        _print_selected_website_column(selection)
        data_rows = rows[1:] if selection.skip_header else rows
        return _load_from_column(data_rows, selection.column_index)
    header_index = _find_header_index(rows[0])
    if header_index is not None:
        return _load_from_column(rows[1:], header_index)
    guess_index = _find_first_website_like_column(rows)
    if guess_index is None:
        return []
    return _load_from_column(rows, guess_index)


def _find_header_index(first_row: list[object]) -> int | None:
    for index, value in enumerate(first_row):
        lowered = str(value or "").strip().lower()
        if lowered in _HEADER_CANDIDATES:
            return index
    return None


def _find_first_website_like_column(rows: list[list[object]]) -> int | None:
    max_width = max((len(row) for row in rows), default=0)
    best_index: int | None = None
    best_score = 0
    for column in range(max_width):
        score = 0
        for row in rows[:50]:
            if column >= len(row):
                continue
            if _normalize_website(str(row[column] or "")):
                score += 1
        if score > best_score:
            best_score = score
            best_index = column
    return best_index if best_score > 0 else None


def _load_from_column(rows: list[list[object]], column: int) -> list[str]:
    result: list[str] = []
    for row in rows:
        if column >= len(row):
            continue
        text = str(row[column] or "").strip()
        if text:
            result.append(text)
    return result


def _pick_website_column(
    rows: list[list[object]],
    *,
    source_name: str,
    website_column_picker: WebsiteColumnPicker | None,
) -> WebsiteColumnSelection | None:
    summaries, has_header = _summarize_columns(rows)
    if not summaries:
        return None
    llm_result = _call_website_column_picker(
        source_name=source_name,
        summaries=summaries,
        website_column_picker=website_column_picker,
    )
    best_summary = _select_best_summary(summaries, llm_result=llm_result)
    if best_summary is None or best_summary.website_count <= 0:
        return None
    confidence = str(llm_result.get("confidence", "") or "").strip().lower() if llm_result else ""
    if confidence not in _WEBSITE_COLUMN_BONUS:
        confidence = "medium" if llm_result else "local"
    reason = _build_selection_reason(best_summary, llm_result=llm_result)
    return WebsiteColumnSelection(
        column_index=best_summary.index,
        header=best_summary.header,
        confidence=confidence,
        reason=reason,
        skip_header=has_header,
    )


def _summarize_columns(rows: list[list[object]]) -> tuple[list[WebsiteColumnSummary], bool]:
    has_header = _looks_like_header_row(rows)
    header_row = rows[0] if has_header else []
    data_rows = rows[1:] if has_header else rows
    max_width = max((len(row) for row in rows), default=0)
    summaries: list[WebsiteColumnSummary] = []
    for index in range(max_width):
        header = str(header_row[index] if index < len(header_row) else f"第{index + 1}列" or "").strip()
        if not header:
            header = f"第{index + 1}列"
        summaries.append(_summarize_single_column(index, header, data_rows))
    return summaries, has_header


def _summarize_single_column(index: int, header: str, rows: list[list[object]]) -> WebsiteColumnSummary:
    samples: list[str] = []
    non_empty_count = 0
    website_count = 0
    homepage_count = 0
    social_count = 0
    email_count = 0
    note_count = 0
    for row in rows[:120]:
        text = _get_cell_text(row, index)
        if not text:
            continue
        non_empty_count += 1
        if len(samples) < 5:
            samples.append(text)
        if _looks_like_email(text):
            email_count += 1
            continue
        normalized = _normalize_website(text)
        if normalized:
            if _is_social_website(normalized):
                social_count += 1
                continue
            website_count += 1
            if _is_homepage_website(normalized):
                homepage_count += 1
            continue
        if _looks_like_note_text(text):
            note_count += 1
    return WebsiteColumnSummary(
        index=index,
        header=header,
        sample_values=samples,
        non_empty_count=non_empty_count,
        website_count=website_count,
        homepage_count=homepage_count,
        social_count=social_count,
        email_count=email_count,
        note_count=note_count,
        local_score=_score_website_column(
            header=header,
            website_count=website_count,
            homepage_count=homepage_count,
            social_count=social_count,
            email_count=email_count,
            note_count=note_count,
            non_empty_count=non_empty_count,
        ),
    )


def _call_website_column_picker(
    *,
    source_name: str,
    summaries: list[WebsiteColumnSummary],
    website_column_picker: WebsiteColumnPicker | None,
) -> dict[str, object]:
    if website_column_picker is None or len(summaries) < 2:
        return {}
    try:
        result = website_column_picker(
            source_name=source_name,
            columns=[summary.to_llm_payload() for summary in summaries],
        )
    except Exception:
        return {}
    return result if isinstance(result, dict) else {}


def _select_best_summary(
    summaries: list[WebsiteColumnSummary],
    *,
    llm_result: dict[str, object],
) -> WebsiteColumnSummary | None:
    selected_index = _coerce_int(llm_result.get("selected_index"))
    confidence = str(llm_result.get("confidence", "") or "").strip().lower()
    llm_bonus = _WEBSITE_COLUMN_BONUS.get(confidence, 45.0 if selected_index is not None else 0.0)
    best_summary: WebsiteColumnSummary | None = None
    best_score = float("-inf")
    for summary in summaries:
        score = summary.local_score
        if selected_index == summary.index:
            score += llm_bonus
        if score > best_score:
            best_score = score
            best_summary = summary
    return best_summary


def _build_selection_reason(summary: WebsiteColumnSummary, *, llm_result: dict[str, object]) -> str:
    selected_index = _coerce_int(llm_result.get("selected_index"))
    llm_reason = str(llm_result.get("reason", "") or "").strip()
    if selected_index == summary.index and llm_reason:
        return llm_reason
    if selected_index is not None and selected_index != summary.index:
        return "本地规则压过了 LLM 误判，当前列更像公司官网列"
    if summary.header.startswith("第"):
        return "该列里官网样本占比最高"
    return "该列表头和样本都更像公司官网列"


def _print_selected_website_column(selection: WebsiteColumnSelection) -> None:
    print(
        f"自动识别网站列：{selection.header} | 置信度: {selection.confidence} | 原因: {selection.reason}",
        flush=True,
    )


def _looks_like_header_row(rows: list[list[object]]) -> bool:
    if not rows:
        return False
    first_row = rows[0]
    if _find_header_index(first_row) is not None:
        return True
    non_empty = [str(value or "").strip() for value in first_row if str(value or "").strip()]
    if not non_empty:
        return False
    if any(_looks_like_email(value) or _normalize_website(value) for value in non_empty):
        return False
    label_like = sum(1 for value in non_empty if _looks_like_header_label(value))
    required = 1 if len(non_empty) == 1 else 2
    return label_like >= max(required, (len(non_empty) + 1) // 2)


def _looks_like_header_label(value: str) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    if _looks_like_email(text) or _normalize_website(text):
        return False
    return len(text) <= 60


def _score_website_column(
    *,
    header: str,
    website_count: int,
    homepage_count: int,
    social_count: int,
    email_count: int,
    note_count: int,
    non_empty_count: int,
) -> float:
    score = _score_header_text(header)
    score += float(website_count * 6)
    score += float(homepage_count * 2)
    score -= float(social_count * 14)
    score -= float(email_count * 16)
    score -= float(note_count * 4)
    if website_count <= 0:
        score -= 40.0
    if social_count > website_count:
        score -= 80.0
    if email_count > website_count:
        score -= 90.0
    if non_empty_count > 0 and website_count / non_empty_count >= 0.6:
        score += 25.0
    return score


def _score_header_text(header: str) -> float:
    lowered = str(header or "").strip().lower()
    score = 0.0
    for hint, value in _POSITIVE_HEADER_HINTS.items():
        if hint in lowered:
            score += float(value)
    for hint, value in _NEGATIVE_HEADER_HINTS.items():
        if hint in lowered:
            score += float(value)
    return score


def _get_cell_text(row: list[object], column: int) -> str:
    if column >= len(row):
        return ""
    return str(row[column] or "").strip()


def _looks_like_email(value: str) -> bool:
    text = str(value or "").strip().lower()
    if not text or " " in text:
        return False
    return "@" in text and "." in text.split("@", 1)[-1]


def _looks_like_note_text(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if len(text) < 24:
        return False
    return " " in text and _normalize_website(text) == "" and not _looks_like_email(text)


def _is_social_website(website: str) -> bool:
    host = str(urlparse(website).netloc or "").strip().lower()
    return any(host == hint or host.endswith(f".{hint}") for hint in _SOCIAL_HOST_HINTS)


def _is_homepage_website(website: str) -> bool:
    path = str(urlparse(website).path or "").strip()
    return path in {"", "/"}


def _coerce_int(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _dedupe_websites(rows: list[str]) -> list[ImportedWebsite]:
    results: list[ImportedWebsite] = []
    seen: set[str] = set()
    for index, raw in enumerate(rows, start=1):
        website = _normalize_website(raw)
        if not website:
            continue
        dedupe_key = _build_dedupe_key(website)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        results.append(
            ImportedWebsite(
                input_index=index,
                raw_website=str(raw).strip(),
                website=website,
                dedupe_key=dedupe_key,
            )
        )
    return results


def _normalize_website(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if _looks_like_email(text):
        return ""
    if "://" not in text:
        text = f"https://{text}"
    parsed = urlparse(text)
    host = str(parsed.netloc or parsed.path or "").strip().lower()
    if not host:
        return ""
    if " " in host or "." not in host:
        return ""
    if host.startswith("www."):
        host = host[4:]
    path = str(parsed.path or "").strip()
    normalized = f"{parsed.scheme or 'https'}://{host}{path}"
    return normalized.rstrip("/") or ""


def _build_dedupe_key(website: str) -> str:
    parsed = urlparse(website)
    host = str(parsed.netloc or "").strip().lower()
    path = str(parsed.path or "").strip()
    if host and path not in {"", "/"}:
        return website.lower()
    if host:
        return host
    return website.lower()
